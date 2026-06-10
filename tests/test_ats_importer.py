from pathlib import Path

from openpyxl import Workbook, load_workbook

from src.ats_importer import (
    AtsHit,
    detect_ats_from_url,
    parse_ats_hits,
    update_company_database,
)


def test_detect_ats_from_common_urls():
    assert detect_ats_from_url("https://boards.greenhouse.io/acme/jobs/123") == (
        "greenhouse",
        "acme",
    )
    assert detect_ats_from_url("https://jobs.lever.co/acme/abc") == (
        "lever",
        "acme",
    )
    assert detect_ats_from_url("https://jobs.ashbyhq.com/acme/123") == (
        "ashby",
        "acme",
    )
    assert detect_ats_from_url("https://acme.wd1.myworkdayjobs.com/External/job/123") == (
        "workday",
        "acme.wd1.myworkdayjobs.com/External",
    )


def test_parse_ats_hits_from_markdown_table():
    markdown = """
| Company | Role | Location | Application |
| --- | --- | --- | --- |
| [Acme Semi](https://example.com) | Hardware Engineer New Grad | Austin, TX | [Apply](https://boards.greenhouse.io/acmesemi/jobs/123) |
| Example Robotics | Firmware Engineer | Boston, MA | [Apply](https://jobs.lever.co/examplerobotics/abc) |
"""

    hits = parse_ats_hits(markdown, "unit")

    assert [hit.company_name for hit in hits] == ["Acme Semi", "Example Robotics"]
    assert hits[0].ats_type == "greenhouse"
    assert hits[0].ats_slug == "acmesemi"
    assert hits[1].ats_type == "lever"
    assert hits[1].ats_slug == "examplerobotics"


def test_parse_ats_hits_from_simplify_html_table_context():
    markdown = """
<td><strong><a href="https://simplify.jobs/c/Further?utm_source=GHList&utm_medium=company">Further</a></strong></td>
<td><div align="center"><a href="https://job-boards.greenhouse.io/furtherearlycareer/jobs/8384012002?utm_source=Simplify&ref=Simplify"><img src="apply.png"></a></div></td>
<td><strong><a href="https://simplify.jobs/c/Authorium?utm_source=GHList&utm_medium=company">Authorium</a></strong></td>
<td><div align="center"><a href="https://jobs.ashbyhq.com/Authorium/e9384068-af40-47b2-83cf-ec76fd8b7222/application?utm_source=Simplify&ref=Simplify"><img src="apply.png"></a></div></td>
"""

    hits = parse_ats_hits(markdown, "unit")

    assert [hit.company_name for hit in hits] == ["Further", "Authorium"]
    assert hits[0].ats_type == "greenhouse"
    assert hits[0].ats_slug == "furtherearlycareer"
    assert hits[1].ats_type == "ashby"
    assert hits[1].ats_slug == "Authorium"


def _create_company_db(path: Path):
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Company_DB"
    headers = [
        "Company Name",
        "Priority",
        "Category",
        "Industry Focus",
        "Target Role Family",
        "Target Keywords",
        "Major Locations",
        "Career Site URL",
        "Monitoring Status",
        "Application Status",
        "Sponsorship Fit",
        "Notes",
        "ATS Type",
        "ATS Company Slug",
        "Official Careers URL",
        "Official Job Search URL",
        "Use Official Scraper",
        "Last Checked At",
        "Last Official Job Count",
    ]
    sheet.append(headers)
    sheet.append(["Acme Semi"] + [""] * (len(headers) - 1))
    workbook.save(path)


def test_update_company_database_existing_company(tmp_path: Path):
    db_path = tmp_path / "company_database.xlsx"
    _create_company_db(db_path)
    hit = AtsHit(
        company_name="Acme Semi",
        ats_type="greenhouse",
        ats_slug="acmesemi",
        official_job_search_url="https://boards.greenhouse.io/acmesemi",
        source_name="unit",
    )

    summary = update_company_database([hit], path=db_path)

    workbook = load_workbook(db_path)
    sheet = workbook["Company_DB"]
    headers = {cell.value: cell.column for cell in sheet[1]}
    assert summary["updated"] == 1
    assert sheet.cell(row=2, column=headers["ATS Type"]).value == "greenhouse"
    assert sheet.cell(row=2, column=headers["ATS Company Slug"]).value == "acmesemi"
    assert sheet.cell(row=2, column=headers["Use Official Scraper"]).value == "Yes"


def test_update_company_database_can_add_missing_company(tmp_path: Path):
    db_path = tmp_path / "company_database.xlsx"
    _create_company_db(db_path)
    hit = AtsHit(
        company_name="New Semi",
        ats_type="lever",
        ats_slug="newsemi",
        official_job_search_url="https://jobs.lever.co/newsemi",
        source_name="unit",
    )

    summary = update_company_database([hit], path=db_path, add_missing=True)

    workbook = load_workbook(db_path)
    sheet = workbook["Company_DB"]
    headers = {cell.value: cell.column for cell in sheet[1]}
    assert summary["added"] == 1
    assert sheet.cell(row=3, column=headers["Company Name"]).value == "New Semi"
    assert sheet.cell(row=3, column=headers["ATS Type"]).value == "lever"


def test_update_company_database_does_not_substring_match_short_names(tmp_path: Path):
    db_path = tmp_path / "company_database.xlsx"
    _create_company_db(db_path)
    workbook = load_workbook(db_path)
    sheet = workbook["Company_DB"]
    sheet.cell(row=3, column=1, value="GM")
    workbook.save(db_path)

    hit = AtsHit(
        company_name="Figma",
        ats_type="greenhouse",
        ats_slug="figma",
        official_job_search_url="https://job-boards.greenhouse.io/figma",
        source_name="unit",
    )

    summary = update_company_database([hit], path=db_path)

    workbook = load_workbook(db_path)
    sheet = workbook["Company_DB"]
    headers = {cell.value: cell.column for cell in sheet[1]}
    assert summary["updated"] == 0
    assert summary["skipped"] == 1
    assert sheet.cell(row=3, column=headers["ATS Type"]).value in {None, ""}
