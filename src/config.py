from __future__ import annotations

from pathlib import Path
from typing import Any
import os

import pandas as pd


BASE_DIR = Path(__file__).resolve().parents[1]
DATA_DIR = BASE_DIR / "data"
COMPANY_MASTER_PATH = DATA_DIR / "company_master.xlsx"
COMPANY_MASTER_NORMALIZED_PATH = DATA_DIR / "company_master_normalized.xlsx"
PENDING_NEW_COMPANIES_PATH = DATA_DIR / "pending_new_companies.xlsx"
COMPANY_DISCOVERY_LOG_PATH = DATA_DIR / "company_discovery_log.csv"
SOURCE_STATE_PATH = DATA_DIR / "source_state.json"
COMPANY_DB_PATH = DATA_DIR / "company_database.xlsx"
SEEN_JOBS_PATH = DATA_DIR / "seen_jobs.json"
JOBS_HISTORY_PATH = DATA_DIR / "jobs_history.csv"

COMPANY_SHEET_NAME = "Company_DB"

COMPANY_COLUMNS = [
    "Company Name",
    "Priority",
    "Category",
    "Industry Focus",
    "Target Role Family",
    "Target Keywords",
    "Major Locations",
    "Career Site URL",
    "Monitoring Status",
    "Application Status",
    "Sponsorship Fit",
    "Notes",
    "ATS Type",
    "ATS Company Slug",
    "Official Careers URL",
    "Official Job Search URL",
    "Use Official Scraper",
    "Last Checked At",
    "Last Official Job Count",
]

DEFAULT_COMPANIES = [
    "Applied Materials",
    "Texas Instruments",
    "Lam Research",
    "KLA",
    "ASML",
    "Micron",
    "Intel",
    "AMD",
    "Qualcomm",
    "Samsung Semiconductor",
    "GlobalFoundries",
    "NXP",
    "Analog Devices",
    "Microchip Technology",
    "Skyworks",
    "Qorvo",
    "Onsemi",
    "Infineon",
    "Keysight",
    "Synopsys",
    "Cadence",
    "Marvell",
    "Broadcom",
    "SLB",
    "Baker Hughes",
    "Emerson",
    "Honeywell",
    "Rockwell Automation",
    "Garmin",
    "Medtronic",
]

DEFAULT_SEARCH_TERMS = [
    "Electrical Engineer New Grad",
    "Hardware Engineer New Grad",
    "Process Engineer New Grad",
    "Process Support Engineer New College Grad",
    "Validation Engineer New Grad",
    "Test Engineer New Grad",
    "Reliability Engineer Entry Level",
    "Failure Analysis Engineer Entry Level",
    "Product Engineer Semiconductor New Grad",
    "Field Applications Engineer New Grad",
    "Customer Engineer Semiconductor New Grad",
    "Applications Engineer Semiconductor New Grad",
    "Firmware Engineer New Grad",
    "Embedded Software Engineer New Grad",
    "Controls Engineer New Grad",
    "Manufacturing Engineer New Grad",
]

DEFAULT_LOCATIONS = [
    "United States",
    "Austin, TX",
    "Houston, TX",
    "Dallas, TX",
    "Phoenix, AZ",
    "Chandler, AZ",
    "Santa Clara, CA",
    "San Jose, CA",
    "Fremont, CA",
    "Hillsboro, OR",
    "Boise, ID",
]

POSITIVE_KEYWORDS = [
    "new grad",
    "new college grad",
    "entry level",
    "electrical engineer",
    "hardware engineer",
    "process engineer",
    "process support engineer",
    "validation engineer",
    "test engineer",
    "reliability engineer",
    "failure analysis",
    "semiconductor",
    "firmware",
    "embedded",
    "field applications",
    "applications engineer",
    "customer engineer",
    "product engineer",
    "manufacturing engineer",
]

NEGATIVE_KEYWORDS = [
    "senior",
    "principal",
    "staff",
    "manager",
    "director",
    "lead engineer",
    "architect",
    "phd required",
    "10+ years",
    "8+ years",
    "7+ years",
    "6+ years",
    "5+ years",
    "4+ years",
    "clearance required",
    "active security clearance",
    "frontend",
    "backend",
    "full stack",
    "web developer",
    "salesforce",
    "devops",
    "cloud infrastructure",
    "engineer iii",
    "engineer iv",
    "postdoc",
    "postdoctoral",
    "civil",
    "highway",
    "structural",
    "forensic",
    "korean bilingual",
]

A_LIST_COMPANIES = {
    name.casefold()
    for name in [
        "Applied Materials",
        "Texas Instruments",
        "Lam Research",
        "KLA",
        "ASML",
        "Micron",
        "Intel",
        "AMD",
        "Qualcomm",
        "Samsung Semiconductor",
        "GlobalFoundries",
        "NXP",
        "Analog Devices",
        "Microchip Technology",
        "Skyworks",
        "Qorvo",
        "Onsemi",
        "Infineon",
        "Keysight",
        "Synopsys",
        "Cadence",
        "Marvell",
        "Broadcom",
    ]
}

SITE_NAMES = ["indeed", "google"]
ENABLE_LINKEDIN = False
RESULTS_WANTED = 20
HOURS_OLD = 72
COUNTRY_INDEED = "USA"
MAX_SEARCH_TERMS = int(os.getenv("MAX_SEARCH_TERMS", "16"))
MAX_LOCATIONS = int(os.getenv("MAX_LOCATIONS", "11"))
MIN_EMAIL_SCORE = int(os.getenv("MIN_EMAIL_SCORE", "6"))
MAX_EMAIL_JOBS = int(os.getenv("MAX_EMAIL_JOBS", "25"))
MIN_COMPANY_EMAIL_SCORE = int(os.getenv("MIN_COMPANY_EMAIL_SCORE", "50"))
MAX_COMPANY_EMAILS = int(os.getenv("MAX_COMPANY_EMAILS", "25"))
AUTO_MERGE_NEW_COMPANIES = os.getenv("AUTO_MERGE_NEW_COMPANIES", "false").casefold() in {
    "1",
    "true",
    "yes",
    "y",
}

COMPANY_MASTER_SHEETS = [
    "Master_Company_List",
    "Company_DB",
    "Apply_Now",
    "Texas_Targets",
    "Strong_H1B",
    "Cold_Targets",
]

DISCOVERY_COMPANY_COLUMNS = [
    "Company Name",
    "Normalized Company Name",
    "Priority",
    "Category",
    "Industry Focus",
    "Target Role Family",
    "Target Keywords",
    "Major Locations",
    "Texas Presence",
    "H1B Sponsor Signal",
    "H1B Evidence URL",
    "Official Careers URL",
    "Official Job Search URL",
    "ATS Type",
    "ATS Company Slug",
    "Use Official Scraper",
    "Monitoring Status",
    "Application Status",
    "Sponsorship Fit",
    "Cold Target Fit",
    "Source",
    "Source URLs",
    "Notes",
    "Last Checked At",
    "Last Updated At",
]


def split_cell(value: Any) -> list[str]:
    if pd.isna(value) or value is None:
        return []
    parts = str(value).replace("\n", ",").replace(";", ",").split(",")
    return [part.strip() for part in parts if part.strip()]


def unique_preserve_order(items: list[str]) -> list[str]:
    seen = set()
    unique = []
    for item in items:
        key = item.casefold()
        if key not in seen:
            unique.append(item)
            seen.add(key)
    return unique


def create_company_database_template(path: Path = COMPANY_DB_PATH) -> None:
    DATA_DIR.mkdir(parents=True, exist_ok=True)
    rows = []
    for company in DEFAULT_COMPANIES:
        rows.append(
            {
                "Company Name": company,
                "Priority": "A" if company.casefold() in A_LIST_COMPANIES else "B",
                "Category": "Target",
                "Industry Focus": "Semiconductor / EE / Hardware",
                "Target Role Family": "Electrical, Hardware, Process, Validation, Firmware, Embedded",
                "Target Keywords": "new grad, entry level, semiconductor",
                "Major Locations": "United States",
                "Career Site URL": "",
                "Monitoring Status": "Active",
                "Application Status": "",
                "Sponsorship Fit": "",
                "Notes": "",
                "ATS Type": "",
                "ATS Company Slug": "",
                "Official Careers URL": "",
                "Official Job Search URL": "",
                "Use Official Scraper": "No",
                "Last Checked At": "",
                "Last Official Job Count": "",
            }
        )
    df = pd.DataFrame(rows, columns=COMPANY_COLUMNS)
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name=COMPANY_SHEET_NAME, index=False)


def ensure_company_database_columns(path: Path = COMPANY_DB_PATH) -> None:
    from openpyxl import load_workbook

    workbook = load_workbook(path)
    if COMPANY_SHEET_NAME not in workbook.sheetnames:
        raise ValueError(f"Missing sheet '{COMPANY_SHEET_NAME}' in {path}")

    sheet = workbook[COMPANY_SHEET_NAME]
    existing_headers = [
        str(cell.value).strip() if cell.value is not None else ""
        for cell in sheet[1]
    ]
    missing_columns = [
        column for column in COMPANY_COLUMNS if column not in existing_headers
    ]
    if not missing_columns:
        header_to_column = {
            str(cell.value).strip(): cell.column
            for cell in sheet[1]
            if cell.value is not None
        }
        old_careers_col = header_to_column.get("Career Site URL")
        official_careers_col = header_to_column.get("Official Careers URL")
        if old_careers_col and official_careers_col:
            changed = False
            for row_index in range(2, sheet.max_row + 1):
                old_value = sheet.cell(row=row_index, column=old_careers_col).value
                official_value = sheet.cell(
                    row=row_index, column=official_careers_col
                ).value
                if old_value and not official_value:
                    sheet.cell(
                        row=row_index, column=official_careers_col, value=old_value
                    )
                    changed = True
            if changed:
                workbook.save(path)
        return

    print(
        "[config] Adding missing company database columns: "
        + ", ".join(missing_columns)
    )
    next_column = sheet.max_column + 1
    for offset, column in enumerate(missing_columns):
        sheet.cell(row=1, column=next_column + offset, value=column)
        if column == "Use Official Scraper":
            for row_index in range(2, sheet.max_row + 1):
                sheet.cell(row=row_index, column=next_column + offset, value="No")

    header_to_column = {
        str(cell.value).strip(): cell.column
        for cell in sheet[1]
        if cell.value is not None
    }
    old_careers_col = header_to_column.get("Career Site URL")
    official_careers_col = header_to_column.get("Official Careers URL")
    if old_careers_col and official_careers_col:
        for row_index in range(2, sheet.max_row + 1):
            old_value = sheet.cell(row=row_index, column=old_careers_col).value
            official_value = sheet.cell(row=row_index, column=official_careers_col).value
            if old_value and not official_value:
                sheet.cell(row=row_index, column=official_careers_col, value=old_value)

    workbook.save(path)


def load_company_database(path: Path = COMPANY_DB_PATH) -> list[dict[str, Any]]:
    if not path.exists():
        print(f"[config] Company database not found. Creating template: {path}")
        create_company_database_template(path)

    ensure_company_database_columns(path)
    df = pd.read_excel(path, sheet_name=COMPANY_SHEET_NAME)
    missing_columns = [column for column in COMPANY_COLUMNS if column not in df.columns]
    if missing_columns:
        missing = ", ".join(missing_columns)
        raise ValueError(
            f"Company database is missing required columns: {missing}. "
            f"Please update {path} or delete it so a fresh template can be created."
        )

    active_df = df[
        df["Monitoring Status"].fillna("").astype(str).str.casefold().ne("paused")
    ].copy()
    return active_df.to_dict(orient="records")


def company_names(companies: list[dict[str, Any]]) -> list[str]:
    return [
        str(company.get("Company Name", "")).strip()
        for company in companies
        if str(company.get("Company Name", "")).strip()
    ]


def build_search_terms(companies: list[dict[str, Any]]) -> list[str]:
    terms = list(DEFAULT_SEARCH_TERMS)
    for company in companies:
        terms.extend(split_cell(company.get("Target Keywords")))
    return unique_preserve_order(terms)[:MAX_SEARCH_TERMS]


def build_locations(companies: list[dict[str, Any]]) -> list[str]:
    locations = list(DEFAULT_LOCATIONS)
    for company in companies:
        locations.extend(split_cell(company.get("Major Locations")))
    return unique_preserve_order(locations)[:MAX_LOCATIONS]
