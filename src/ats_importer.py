from __future__ import annotations

import argparse
import re
from dataclasses import dataclass
from datetime import datetime, timezone
from pathlib import Path
from typing import Iterable
from urllib.parse import urlparse

import requests
from openpyxl import load_workbook

from src import config


SIMPLIFY_SOURCES = {
    "new-grad": "https://raw.githubusercontent.com/SimplifyJobs/New-Grad-Positions/dev/README.md",
    "internships": "https://raw.githubusercontent.com/SimplifyJobs/Summer2026-Internships/dev/README.md",
}

ATS_COLUMNS = [
    "ATS Type",
    "ATS Company Slug",
    "Official Careers URL",
    "Official Job Search URL",
    "Use Official Scraper",
    "Last Checked At",
]

REQUEST_TIMEOUT_SECONDS = 30
SUPPORTED_ACTIVE_ATS_TYPES = {"greenhouse", "lever", "ashby"}


@dataclass(frozen=True)
class AtsHit:
    company_name: str
    ats_type: str
    ats_slug: str
    official_job_search_url: str
    source_name: str


def normalize_company_name(name: str) -> str:
    normalized = name.casefold()
    normalized = re.sub(r"\([^)]*\)", " ", normalized)
    normalized = re.sub(r"\b(inc|inc\.|corp|corp\.|corporation|co|co\.|company|ltd|llc|plc|technology|technologies)\b", " ", normalized)
    normalized = re.sub(r"[^a-z0-9]+", " ", normalized)
    return " ".join(normalized.split())


def clean_markdown_text(text: str) -> str:
    text = re.sub(r"!\[[^\]]*\]\([^)]*\)", "", text)
    text = re.sub(r"\[([^\]]+)\]\([^)]*\)", r"\1", text)
    text = re.sub(r"<[^>]+>", "", text)
    text = text.replace("↳", "").replace("✅", "").replace("🛂", "")
    text = re.sub(r"\s+", " ", text)
    return text.strip()


def extract_urls(text: str) -> list[str]:
    urls = re.findall(r"https?://[^\s)>\"]+", text)
    return [url.rstrip(".,|") for url in urls]


def extract_company_from_markdown_line(line: str) -> str:
    if "|" in line:
        cells = [cell.strip() for cell in line.strip().strip("|").split("|")]
        if cells and cells[0] and not cells[0].casefold().startswith("company"):
            return clean_markdown_text(cells[0])
    simplify_company = re.search(
        r'<a\s+href="https://simplify\.jobs/c/[^"]+"[^>]*>(.*?)</a>',
        line,
        re.IGNORECASE,
    )
    if simplify_company:
        return clean_markdown_text(simplify_company.group(1))
    return ""


def _slug_from_path(path: str, marker: str | None = None) -> str:
    parts = [part for part in path.split("/") if part]
    if marker and marker in parts:
        marker_index = parts.index(marker)
        if len(parts) > marker_index + 1:
            return parts[marker_index + 1]
    return parts[0] if parts else ""


def detect_ats_from_url(url: str) -> tuple[str, str] | None:
    parsed = urlparse(url)
    host = parsed.netloc.casefold()
    path = parsed.path

    if "greenhouse.io" in host:
        if "boards-api.greenhouse.io" in host:
            slug = _slug_from_path(path, "boards")
        else:
            slug = _slug_from_path(path)
        if slug in {"embed", "job_app"}:
            return None
        return ("greenhouse", slug) if slug else None

    if "lever.co" in host:
        if host.startswith("api.") or "api.eu.lever.co" in host:
            slug = _slug_from_path(path, "postings")
        else:
            slug = _slug_from_path(path)
        return ("lever", slug) if slug else None

    if "ashbyhq.com" in host:
        if "api.ashbyhq.com" in host:
            slug = _slug_from_path(path, "job-board")
        else:
            slug = _slug_from_path(path)
        return ("ashby", slug) if slug else None

    if "myworkdayjobs.com" in host:
        site_slug = _slug_from_path(path)
        slug = f"{parsed.netloc}/{site_slug}" if site_slug else parsed.netloc
        return ("workday", slug)

    return None


def parse_ats_hits(markdown_text: str, source_name: str) -> list[AtsHit]:
    hits: list[AtsHit] = []
    seen = set()
    current_company = ""
    for line in markdown_text.splitlines():
        company_name = extract_company_from_markdown_line(line)
        if company_name:
            current_company = company_name
        elif current_company:
            company_name = current_company
        else:
            continue

        for url in extract_urls(line):
            ats = detect_ats_from_url(url)
            if not ats:
                continue
            ats_type, ats_slug = ats
            key = (
                normalize_company_name(company_name),
                ats_type,
                ats_slug.casefold(),
            )
            if key in seen:
                continue
            seen.add(key)
            hits.append(
                AtsHit(
                    company_name=company_name,
                    ats_type=ats_type,
                    ats_slug=ats_slug,
                    official_job_search_url=url,
                    source_name=source_name,
                )
            )
    return hits


def fetch_source(source_name: str, url: str) -> str:
    print(f"[ats_importer] Fetching {source_name}: {url}")
    response = requests.get(url, timeout=REQUEST_TIMEOUT_SECONDS)
    response.raise_for_status()
    return response.text


def collect_hits(source_names: Iterable[str]) -> list[AtsHit]:
    hits: list[AtsHit] = []
    for source_name in source_names:
        source_url = SIMPLIFY_SOURCES[source_name]
        markdown = fetch_source(source_name, source_url)
        source_hits = parse_ats_hits(markdown, source_name)
        print(f"[ats_importer] {source_name}: {len(source_hits)} ATS links found")
        hits.extend(source_hits)
    return hits


def _header_map(sheet) -> dict[str, int]:
    return {
        str(cell.value).strip(): cell.column
        for cell in sheet[1]
        if cell.value is not None
    }


def _find_company_row(sheet, headers: dict[str, int], company_name: str) -> int | None:
    target = normalize_company_name(company_name)
    company_col = headers["Company Name"]
    for row_index in range(2, sheet.max_row + 1):
        value = sheet.cell(row=row_index, column=company_col).value
        if normalize_company_name(str(value or "")) == target:
            return row_index
    return None


def _append_company_row(sheet, headers: dict[str, int], hit: AtsHit) -> int:
    row_index = sheet.max_row + 1
    defaults = {
        "Company Name": hit.company_name,
        "Priority": "C",
        "Category": "Imported",
        "Industry Focus": "Imported from SimplifyJobs",
        "Target Role Family": "Electrical, Hardware, Process, Validation, Firmware, Embedded",
        "Target Keywords": "new grad, entry level, hardware, electrical, semiconductor",
        "Major Locations": "United States",
        "Monitoring Status": "Watchlist",
        "Application Status": "",
        "Sponsorship Fit": "",
        "Notes": f"Imported ATS link from {hit.source_name}",
    }
    for column_name, value in defaults.items():
        if column_name in headers:
            sheet.cell(row=row_index, column=headers[column_name], value=value)
    return row_index


def update_company_database(
    hits: list[AtsHit],
    path: Path = config.COMPANY_DB_PATH,
    add_missing: bool = False,
    dry_run: bool = False,
    limit: int | None = None,
) -> dict[str, int]:
    if limit is not None:
        hits = hits[:limit]

    config.ensure_company_database_columns(path)
    workbook = load_workbook(path)
    sheet = workbook[config.COMPANY_SHEET_NAME]
    headers = _header_map(sheet)
    missing_headers = [column for column in config.COMPANY_COLUMNS if column not in headers]
    if missing_headers:
        raise ValueError(f"Missing required columns after migration: {missing_headers}")

    updated = 0
    added = 0
    skipped = 0
    seen_company_ats = set()
    checked_at = datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M:%S UTC")

    for hit in hits:
        key = (normalize_company_name(hit.company_name), hit.ats_type, hit.ats_slug.casefold())
        if key in seen_company_ats:
            continue
        seen_company_ats.add(key)

        row_index = _find_company_row(sheet, headers, hit.company_name)
        if row_index is None:
            if not add_missing:
                skipped += 1
                continue
            row_index = _append_company_row(sheet, headers, hit)
            added += 1

        if dry_run:
            updated += 1
            continue

        sheet.cell(row=row_index, column=headers["ATS Type"], value=hit.ats_type)
        sheet.cell(row=row_index, column=headers["ATS Company Slug"], value=hit.ats_slug)
        sheet.cell(
            row=row_index,
            column=headers["Official Job Search URL"],
            value=hit.official_job_search_url,
        )
        if not sheet.cell(row=row_index, column=headers["Official Careers URL"]).value:
            sheet.cell(
                row=row_index,
                column=headers["Official Careers URL"],
                value=hit.official_job_search_url,
            )
        use_official = "Yes" if hit.ats_type in SUPPORTED_ACTIVE_ATS_TYPES else "No"
        sheet.cell(row=row_index, column=headers["Use Official Scraper"], value=use_official)
        sheet.cell(row=row_index, column=headers["Last Checked At"], value=checked_at)
        note_cell = sheet.cell(row=row_index, column=headers["Notes"])
        existing_note = str(note_cell.value or "").strip()
        if hit.ats_type in SUPPORTED_ACTIVE_ATS_TYPES:
            import_note = f"ATS imported from {hit.source_name}"
        else:
            import_note = f"ATS imported from {hit.source_name}; scraper not active for {hit.ats_type}"
        if import_note not in existing_note:
            note_cell.value = f"{existing_note}; {import_note}".strip("; ")
        updated += 1

    if not dry_run:
        workbook.save(path)

    return {"updated": updated, "added": added, "skipped": skipped, "hits": len(hits)}


def clear_imported_ats_settings(path: Path = config.COMPANY_DB_PATH) -> int:
    config.ensure_company_database_columns(path)
    workbook = load_workbook(path)
    sheet = workbook[config.COMPANY_SHEET_NAME]
    headers = _header_map(sheet)
    cleared = 0

    for row_index in range(2, sheet.max_row + 1):
        notes = str(sheet.cell(row=row_index, column=headers["Notes"]).value or "")
        if "ATS imported from" not in notes:
            continue
        for column_name in ATS_COLUMNS:
            sheet.cell(row=row_index, column=headers[column_name], value="")
        notes = re.sub(r";?\s*ATS imported from [^;]+", "", notes).strip("; ")
        sheet.cell(row=row_index, column=headers["Notes"], value=notes)
        cleared += 1

    workbook.save(path)
    return cleared


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Import ATS settings from public job-list repos.")
    parser.add_argument(
        "--source",
        choices=["all", *SIMPLIFY_SOURCES.keys()],
        default="all",
        help="Which source to import from.",
    )
    parser.add_argument(
        "--add-missing",
        action="store_true",
        help="Add companies that are not already in company_database.xlsx.",
    )
    parser.add_argument(
        "--dry-run",
        action="store_true",
        help="Parse and match without saving the workbook.",
    )
    parser.add_argument(
        "--limit",
        type=int,
        default=None,
        help="Limit number of parsed ATS hits processed, useful for testing.",
    )
    parser.add_argument(
        "--clear-imported",
        action="store_true",
        help="Clear ATS fields previously written by this importer.",
    )
    return parser.parse_args()


def main() -> dict[str, int]:
    args = parse_args()
    if args.clear_imported:
        cleared = clear_imported_ats_settings()
        print(f"[ats_importer] Cleared imported ATS settings from {cleared} rows")
        return {"cleared": cleared}

    source_names = list(SIMPLIFY_SOURCES) if args.source == "all" else [args.source]
    hits = collect_hits(source_names)
    summary = update_company_database(
        hits,
        add_missing=args.add_missing,
        dry_run=args.dry_run,
        limit=args.limit,
    )
    mode = "DRY RUN" if args.dry_run else "SAVED"
    print(
        f"[ats_importer] {mode}: hits={summary['hits']} "
        f"updated={summary['updated']} added={summary['added']} skipped={summary['skipped']}"
    )
    return summary


if __name__ == "__main__":
    main()
